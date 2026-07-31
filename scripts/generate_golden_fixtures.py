#!/usr/bin/env python3
"""Synthetic Golden Test Fixture Generator for Finance Intelligence.

Generates synthetic PDF, XLSX, and CSV golden test files for Phase 2 parser testing.
All generated files are strictly non-sensitive synthetic test data and documented as
SYNTHETIC_TEST_DATA.
"""

import hashlib
import json
import os

import openpyxl
from pypdf import PdfWriter

FIXTURE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "tests", "fixtures", "golden"))


def create_synthetic_pdf(filepath: str):
    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    # Note: Blank PDF page with metadata

    with open(filepath, "wb") as f:
        writer.write(f)


def create_encrypted_pdf(filepath: str):
    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    writer.encrypt("test_password_123")
    with open(filepath, "wb") as f:
        writer.write(f)


def create_malformed_pdf(filepath: str):
    with open(filepath, "wb") as f:
        f.write(b"%PDF-1.4\n%TRUNCATED_INVALID_HEADER\n")


def create_minimal_xlsx(filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Summary"
    ws.append(["Category", "Q1 Amount", "Q2 Amount", "Status"])
    ws.append(["Revenue", 150000.0, 180000.0, "Audited"])
    ws.append(["Operating Expenses", 45000.0, 52000.0, "Audited"])
    ws.append(["Net Income", "=B3-B4", "=C3-C4", "Calculated"])
    ws.merge_cells("A1:D1")

    ws_hidden = wb.create_sheet(title="Internal Notes")
    ws_hidden.append(["Note ID", "Comment"])
    ws_hidden.append([1, "Confidential adjustment"])
    ws_hidden.sheet_state = "hidden"

    wb.save(filepath)


def create_formula_injection_csv(filepath: str):
    content = (
        "Account,Amount,Comment\n"
        "Sales,10000,Standard Sale\n"
        "=1+2,5000,Formula Injection Test\n"
        "@SUM(1..10),2500,At Symbol Injection\n"
        "-500+2,1200,Minus Prefix Test\n"
        "+1000,3400,Plus Prefix Test\n"
    )
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)


def create_turkish_csv(filepath: str):
    content = (
        "Şehir,Gelir,Açıklama\n"
        "İstanbul,250000.50,Şirket Ana Merkezi\n"
        "İzmir,120000.00,Ege Bölge Müdürlüğü\n"
        "Eskişehir,85000.75,Şube Satışları\n"
    )
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)


def create_bom_csv(filepath: str):
    content = "Header1,Header2\nValue1,Value2\n"
    with open(filepath, "w", encoding="utf-8-sig") as f:
        f.write(content)


def generate_fixtures():
    os.makedirs(FIXTURE_DIR, exist_ok=True)

    files = {
        "sample_financial.pdf": lambda p: create_synthetic_pdf(p),
        "sample_encrypted.pdf": lambda p: create_encrypted_pdf(p),
        "sample_malformed.pdf": lambda p: create_malformed_pdf(p),
        "sample_ledger.xlsx": lambda p: create_minimal_xlsx(p),
        "sample_formula_injection.csv": lambda p: create_formula_injection_csv(p),
        "sample_turkish.csv": lambda p: create_turkish_csv(p),
        "sample_bom.csv": lambda p: create_bom_csv(p),
    }

    manifest = {"dataset": "SYNTHETIC_TEST_DATA", "fixtures": {}}

    for fname, func in files.items():
        fpath = os.path.join(FIXTURE_DIR, fname)
        func(fpath)
        with open(fpath, "rb") as f:
            h = hashlib.sha256(f.read()).hexdigest()
        manifest["fixtures"][fname] = {
            "sha256": h,
            "size_bytes": os.path.getsize(fpath),
            "classification": "SYNTHETIC_TEST_DATA",
        }

    manifest_path = os.path.join(FIXTURE_DIR, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    print(f"Generated {len(files)} synthetic golden fixtures and manifest in {FIXTURE_DIR}")


if __name__ == "__main__":
    generate_fixtures()
