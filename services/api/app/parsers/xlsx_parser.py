import zipfile
from io import BytesIO

import openpyxl
from app.core.config import settings
from app.parsers.base import (
    CanonicalExtractionOutput,
    DocumentParserPort,
    ExtractionWarningItem,
)
from openpyxl.utils.exceptions import InvalidFileException

FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


class XlsxParser(DocumentParserPort):
    PARSER_NAME = "XlsxParser"
    PARSER_VERSION = "1.0.0"

    def parse(self, file_bytes: bytes, file_name: str) -> CanonicalExtractionOutput:
        # ZIP Archive & ZIP Bomb Safety Inspection
        try:
            with zipfile.ZipFile(BytesIO(file_bytes), "r") as zf:
                infolist = zf.infolist()
                if len(infolist) > settings.MAX_ZIP_ENTRIES:
                    return CanonicalExtractionOutput(
                        parser_name=self.PARSER_NAME,
                        parser_version=self.PARSER_VERSION,
                        status="REJECTED",
                        quality_score=0.0,
                        text_layer_present=False,
                        pages=[],
                        chunks=[],
                        warnings=[
                            ExtractionWarningItem(
                                warning_code="ZIP_BOMB_LIMIT_EXCEEDED",
                                warning_message=f"XLSX archive contains {len(infolist)} ZIP entries, exceeding limit of {settings.MAX_ZIP_ENTRIES}.",
                                lineage_ref={"format": "xlsx", "zip_entries": len(infolist)},
                            )
                        ],
                    )

                total_uncompressed = 0
                for info in infolist:
                    if info.file_size > settings.MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES:
                        return CanonicalExtractionOutput(
                            parser_name=self.PARSER_NAME,
                            parser_version=self.PARSER_VERSION,
                            status="REJECTED",
                            quality_score=0.0,
                            text_layer_present=False,
                            pages=[],
                            chunks=[],
                            warnings=[
                                ExtractionWarningItem(
                                    warning_code="ZIP_BOMB_LIMIT_EXCEEDED",
                                    warning_message=f"XLSX ZIP entry '{info.filename}' uncompressed size ({info.file_size} bytes) exceeds limit of {settings.MAX_ZIP_ENTRY_UNCOMPRESSED_BYTES}.",
                                    lineage_ref={"format": "xlsx", "entry_filename": info.filename},
                                )
                            ],
                        )
                    total_uncompressed += info.file_size
                    if info.compress_size > 0:
                        ratio = info.file_size / float(info.compress_size)
                        if ratio > settings.MAX_ZIP_COMPRESSION_RATIO and info.file_size > 1024:
                            return CanonicalExtractionOutput(
                                parser_name=self.PARSER_NAME,
                                parser_version=self.PARSER_VERSION,
                                status="REJECTED",
                                quality_score=0.0,
                                text_layer_present=False,
                                pages=[],
                                chunks=[],
                                warnings=[
                                    ExtractionWarningItem(
                                        warning_code="ZIP_BOMB_LIMIT_EXCEEDED",
                                        warning_message=f"XLSX ZIP entry '{info.filename}' compression ratio ({ratio:.1f}) exceeds limit of {settings.MAX_ZIP_COMPRESSION_RATIO}.",
                                        lineage_ref={"format": "xlsx", "entry_filename": info.filename, "ratio": ratio},
                                    )
                                ],
                            )

                if total_uncompressed > settings.MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES:
                    return CanonicalExtractionOutput(
                        parser_name=self.PARSER_NAME,
                        parser_version=self.PARSER_VERSION,
                        status="REJECTED",
                        quality_score=0.0,
                        text_layer_present=False,
                        pages=[],
                        chunks=[],
                        warnings=[
                            ExtractionWarningItem(
                                warning_code="ZIP_BOMB_LIMIT_EXCEEDED",
                                warning_message=f"XLSX total uncompressed size ({total_uncompressed} bytes) exceeds limit of {settings.MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES}.",
                                lineage_ref={"format": "xlsx", "total_uncompressed": total_uncompressed},
                            )
                        ],
                    )
        except zipfile.BadZipFile:
            pass

        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
        except (zipfile.BadZipFile, InvalidFileException, KeyError, OSError) as err:
            return CanonicalExtractionOutput(
                parser_name=self.PARSER_NAME,
                parser_version=self.PARSER_VERSION,
                status="FAILED",
                quality_score=0.0,
                text_layer_present=False,
                pages=[],
                chunks=[],
                warnings=[
                    ExtractionWarningItem(
                        warning_code="MALFORMED_DOCUMENT",
                        warning_message=f"XLSX container is corrupted or invalid: {err!s}",
                        lineage_ref={"format": "xlsx"},
                    )
                ],
            )

        chunks_output = []
        warnings = []
        chunk_index = 0
        total_extracted_chars = 0

        max_sheets = settings.MAX_XLSX_SHEETS
        max_rows = settings.MAX_XLSX_ROWS
        max_cols = settings.MAX_XLSX_COLS
        max_cell_len = settings.MAX_CELL_LEN
        max_chars = settings.MAX_EXTRACTED_CHARS

        sheet_count = 0
        for sheet_name in wb.sheetnames:
            sheet_count += 1
            if sheet_count > max_sheets:
                warnings.append(
                    ExtractionWarningItem(
                        warning_code="RESOURCE_LIMIT_EXCEEDED",
                        warning_message=f"XLSX sheet count exceeds safety limit of {max_sheets}.",
                        lineage_ref={"format": "xlsx", "sheet_name": sheet_name},
                    )
                )
                break

            ws = wb[sheet_name]
            row_idx = 0
            headers: list[str] = []

            for row in ws.iter_rows(values_only=True):
                row_idx += 1
                if row_idx > max_rows:
                    warnings.append(
                        ExtractionWarningItem(
                            warning_code="RESOURCE_LIMIT_EXCEEDED",
                            warning_message=f"XLSX sheet '{sheet_name}' row count exceeds limit of {max_rows}.",
                            lineage_ref={"format": "xlsx", "sheet_name": sheet_name, "row_index": row_idx},
                        )
                    )
                    break

                if len(row) > max_cols:
                    warnings.append(
                        ExtractionWarningItem(
                            warning_code="RESOURCE_LIMIT_EXCEEDED",
                            warning_message=f"XLSX sheet '{sheet_name}' column count ({len(row)}) exceeds limit of {max_cols}.",
                            lineage_ref={"format": "xlsx", "sheet_name": sheet_name, "row_index": row_idx},
                        )
                    )
                    break

                row_str_values = [str(val) if val is not None else "" for val in row]

                if row_idx == 1 and not headers:
                    headers = row_str_values
                    chunk_index += 1
                    header_content = " | ".join(headers)
                    total_extracted_chars += len(header_content)
                    chunks_output.append(
                        {
                            "chunk_index": chunk_index,
                            "chunk_type": "HEADER",
                            "content": header_content,
                            "source_lineage": {
                                "format": "xlsx",
                                "sheet_name": sheet_name,
                                "row_index": row_idx,
                                "is_header": True,
                            },
                        }
                    )
                    continue

                for col_idx, cell_val in enumerate(row_str_values):
                    if not cell_val:
                        continue

                    raw_value = cell_val
                    if len(raw_value) > max_cell_len:
                        raw_value = raw_value[:max_cell_len]
                        warnings.append(
                            ExtractionWarningItem(
                                warning_code="RESOURCE_LIMIT_EXCEEDED",
                                warning_message=f"Cell length in sheet '{sheet_name}' truncated to max length {max_cell_len}.",
                                lineage_ref={"format": "xlsx", "sheet_name": sheet_name, "row_index": row_idx},
                            )
                        )

                    has_risk = any(raw_value.startswith(prefix) for prefix in FORMULA_PREFIXES)
                    is_formula = raw_value.startswith("=")
                    leading_char = raw_value[0] if (has_risk and len(raw_value) > 0) else None

                    if has_risk:
                        warnings.append(
                            ExtractionWarningItem(
                                warning_code="XLSX_FORMULA_INJECTION_RISK",
                                warning_message=f"Formula injection risk in sheet '{sheet_name}', row {row_idx}, col {col_idx + 1}: leading '{leading_char}'",
                                lineage_ref={
                                    "format": "xlsx",
                                    "sheet_name": sheet_name,
                                    "row_index": row_idx,
                                    "column_index": col_idx + 1,
                                    "leading_char": leading_char,
                                },
                            )
                        )

                    header_name = headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx + 1}"

                    chunk_index += 1
                    total_extracted_chars += len(raw_value)

                    lineage_dict = {
                        "format": "xlsx",
                        "sheet_name": sheet_name,
                        "row_index": row_idx,
                        "column_index": col_idx + 1,
                        "column_name": header_name,
                        "is_formula": is_formula,
                        "formula_injection_risk": has_risk,
                        "leading_formula_character": leading_char,
                    }
                    if is_formula:
                        lineage_dict["formula_text"] = raw_value

                    chunks_output.append(
                        {
                            "chunk_index": chunk_index,
                            "chunk_type": "CELL",
                            "content": raw_value,
                            "source_lineage": lineage_dict,
                        }
                    )

                    if total_extracted_chars > max_chars:
                        warnings.append(
                            ExtractionWarningItem(
                                warning_code="RESOURCE_LIMIT_EXCEEDED",
                                warning_message=f"Extracted character limit ({max_chars}) reached.",
                                lineage_ref={"format": "xlsx"},
                            )
                        )
                        break

        wb.close()
        status = "COMPLETED_WITH_WARNINGS" if warnings else "COMPLETED"

        quality_score = 0.90 if warnings else 1.0

        return CanonicalExtractionOutput(
            parser_name=self.PARSER_NAME,
            parser_version=self.PARSER_VERSION,
            status=status,
            quality_score=quality_score,
            text_layer_present=True,
            pages=[],
            chunks=chunks_output,
            warnings=warnings,
        )
